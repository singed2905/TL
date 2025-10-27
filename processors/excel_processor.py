import pandas as pd
import json
import os
import openpyxl
from typing import Dict, List, Tuple, Any, Optional
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import re


class ExcelProcessor:
    def __init__(self, mapping_file: str = "config/excel_mapping.json"):
        self.mapping_file = mapping_file
        self.mapping = self._load_mapping()

    def _load_mapping(self) -> Dict:
        """Load mapping configuration from JSON file"""
        try:
            with open(self.mapping_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Không thể load file mapping: {str(e)}")
            return self._get_default_mapping()

    def _get_default_mapping(self) -> Dict:
        """Tạo mapping mặc định nếu không load được file"""
        return {
            "equation_mapping_by_phuong_trinh": {
                "2_an": {
                    "required_columns": ["Phương trình 1", "Phương trình 2"],
                    "columns": {
                        "phuong_trinh_1": {"excel_column": "Phương trình 1"},
                        "phuong_trinh_2": {"excel_column": "Phương trình 2"}
                    }
                },
                "3_an": {
                    "required_columns": ["Phương trình 1", "Phương trình 2", "Phương trình 3"],
                    "columns": {
                        "phuong_trinh_1": {"excel_column": "Phương trình 1"},
                        "phuong_trinh_2": {"excel_column": "Phương trình 2"},
                        "phuong_trinh_3": {"excel_column": "Phương trình 3"}
                    }
                },
                "4_an": {
                    "required_columns": ["Phương trình 1", "Phương trình 2", "Phương trình 3", "Phương trình 4"],
                    "columns": {
                        "phuong_trinh_1": {"excel_column": "Phương trình 1"},
                        "phuong_trinh_2": {"excel_column": "Phương trình 2"},
                        "phuong_trinh_3": {"excel_column": "Phương trình 3"},
                        "phuong_trinh_4": {"excel_column": "Phương trình 4"}
                    }
                }
            }
        }

    def validate_excel_structure(self, df: pd.DataFrame, shape_a: str, shape_b: str = None) -> Tuple[bool, List[str]]:
        """Validate Excel structure against selected shapes"""
        missing_columns = []

        # Check Group A columns
        if shape_a in self.mapping['group_a_mapping']:
            required_cols = self.mapping['group_a_mapping'][shape_a]['required_columns']
            for col in required_cols:
                if col not in df.columns:
                    missing_columns.append(f"Nhóm A - {col}")

        # Check Group B columns
        if shape_b and shape_b in self.mapping['group_b_mapping']:
            required_cols = self.mapping['group_b_mapping'][shape_b]['required_columns']
            for col in required_cols:
                if col not in df.columns:
                    missing_columns.append(f"Nhóm B - {col}")

        return len(missing_columns) == 0, missing_columns

    def read_excel_data(self, file_path: str) -> pd.DataFrame:
        """Read Excel file and normalize data"""
        try:
            df = pd.read_excel(file_path)
            # Normalize column names (remove extra spaces)
            df.columns = df.columns.str.strip()
            return df
        except Exception as e:
            raise Exception(f"Không thể đọc file Excel: {str(e)}")

    def extract_shape_data(self, row: pd.Series, shape_type: str, group: str) -> Dict:
        """Extract data for specific shape from row"""
        if group == 'A':
            shape_mapping = self.mapping['group_a_mapping'].get(shape_type, {})
        else:
            shape_mapping = self.mapping['group_b_mapping'].get(shape_type, {})

        data_dict = {}
        for field, config in shape_mapping.get('columns', {}).items():
            excel_column = config.get('excel_column')
            if excel_column and excel_column in row:
                data_dict[field] = str(row[excel_column]).strip()

        return data_dict

    def validate_equation_structure_by_phuong_trinh(self, df: pd.DataFrame, so_an: int) -> Tuple[bool, List[str]]:
        """Validate Excel structure for equation mode by phuong trinh"""
        missing_columns = []
        key = f"{so_an}_an"

        if key not in self.mapping['equation_mapping_by_phuong_trinh']:
            return False, [f"Không tìm thấy cấu hình cho hệ {so_an} ẩn"]

        required_cols = self.mapping['equation_mapping_by_phuong_trinh'][key]['required_columns']
        for col in required_cols:
            if col not in df.columns:
                missing_columns.append(col)

        return len(missing_columns) == 0, missing_columns

    def extract_equation_coefficients_by_phuong_trinh(self, df: pd.DataFrame, so_an: int) -> List[str]:
        """Extract equation coefficients from Excel by phuong trinh - chỉ dòng đầu tiên"""
        return self._extract_single_row_coefficients(df, so_an, 0)

    def _extract_single_row_coefficients(self, df: pd.DataFrame, so_an: int, row_index: int = 0) -> List[str]:
        """
        Trích xuất hệ số từ một dòng cụ thể, đảm bảo mỗi phương trình đủ (so_an + 1) hệ số (nếu thiếu thì bù 0),
        không bao giờ để trộn block hệ số giữa các phương trình khác nhau.
        """
        key = f"{so_an}_an"
        if key not in self.mapping['equation_mapping_by_phuong_trinh']:
            return []

        columns_config = self.mapping['equation_mapping_by_phuong_trinh'][key]['columns']
        he_so_list = []

        # Chỉ lấy dòng được chỉ định
        if row_index >= len(df):
            return []

        row = df.iloc[row_index]
        num_coeff_per_eq = so_an + 1  # VD: 3 ẩn = 4 hệ số/phương trình

        for col_config in columns_config.values():
            excel_column = col_config['excel_column']
            coefficients = []
            if excel_column in row:
                value = row[excel_column]
                if pd.isna(value):
                    value = ""
                coefficients = [coeff.strip() for coeff in str(value).strip().split(',') if coeff.strip() != ""]
            # Bổ sung số 0 cho đủ mỗi phương trình
            while len(coefficients) < num_coeff_per_eq:
                coefficients.append("0")
            coefficients = coefficients[:num_coeff_per_eq]  # Giới hạn đúng số cần
            he_so_list.extend(coefficients)

        return he_so_list

    def extract_all_equation_rows(self, df: pd.DataFrame, so_an: int) -> List[Dict[str, Any]]:
        """
        Trích xuất tất cả các dòng hệ số từ DataFrame - PHIÊN BẢN ĐÃ SỬA

        Returns:
            List of dictionaries, mỗi dict chứa:
            - 'row_index': chỉ số dòng (bắt đầu từ 0)
            - 'he_so': danh sách hệ số
            - 'original_data': dữ liệu gốc từ dòng đó
            - 'has_data': có dữ liệu hợp lệ không
        """
        key = f"{so_an}_an"
        if key not in self.mapping['equation_mapping_by_phuong_trinh']:
            return []

        columns_config = self.mapping['equation_mapping_by_phuong_trinh'][key]['columns']
        required_cols = self.mapping['equation_mapping_by_phuong_trinh'][key]['required_columns']

        all_rows_data = []
        num_coeff_per_eq = so_an + 1  # Số hệ số mỗi phương trình

        for row_index in range(len(df)):
            row_data = {
                'row_index': row_index,
                'he_so': [],
                'original_data': {},
                'has_data': False
            }

            # Lưu dữ liệu gốc
            for col in required_cols:
                if col in df.columns:
                    value = df.iloc[row_index][col]
                    row_data['original_data'][col] = value if not pd.isna(value) else ""

            # Trích xuất hệ số - SỬA QUAN TRỌNG Ở ĐÂY
            he_so_list = []
            for col_config in columns_config.values():
                excel_column = col_config['excel_column']
                coefficients = []

                if excel_column in df.columns:
                    value = df.iloc[row_index][excel_column]
                    if pd.isna(value):
                        value = ""
                    # Tách hệ số và làm sạch
                    raw_coefficients = [coeff.strip() for coeff in str(value).strip().split(',') if coeff.strip() != ""]

                    # Xử lý từng phương trình RIÊNG BIỆT
                    coefficients = raw_coefficients.copy()

                    # Bổ sung số 0 cho đủ mỗi phương trình
                    while len(coefficients) < num_coeff_per_eq:
                        coefficients.append("0")

                    # Giới hạn đúng số hệ số cần
                    coefficients = coefficients[:num_coeff_per_eq]

                else:
                    # Nếu không có cột, tạo danh sách toàn 0
                    coefficients = ["0"] * num_coeff_per_eq

                # Thêm hệ số của phương trình này vào danh sách
                he_so_list.extend(coefficients)

            # Kiểm tra xem dòng có dữ liệu không
            has_valid_data = any(
                str(row_data['original_data'].get(col, "")).strip() not in ["", "0"]
                for col in required_cols
                if col in df.columns
            )

            row_data['he_so'] = he_so_list
            row_data['has_data'] = has_valid_data
            all_rows_data.append(row_data)

        return all_rows_data

    def process_equation_batch(self, file_path: str, so_an: int) -> List[Dict[str, Any]]:
        """
        Xử lý hàng loạt các hệ phương trình từ Excel file

        Args:
            file_path: Đường dẫn đến file Excel
            so_an: Số ẩn của hệ phương trình

        Returns:
            Danh sách các dictionary chứa dữ liệu từng dòng
        """
        try:
            # Đọc file Excel
            df = self.read_excel_data(file_path)

            # Validate cấu trúc
            is_valid, missing_cols = self.validate_equation_structure_by_phuong_trinh(df, so_an)
            if not is_valid:
                raise Exception(f"File Excel thiếu các cột bắt buộc: {', '.join(missing_cols)}")

            # Trích xuất tất cả các dòng
            all_rows = self.extract_all_equation_rows(df, so_an)

            # Lọc các dòng có dữ liệu hợp lệ
            valid_rows = [row for row in all_rows if row['has_data']]

            if not valid_rows:
                raise Exception("Không tìm thấy dòng nào có dữ liệu hợp lệ")

            return valid_rows

        except Exception as e:
            raise Exception(f"Lỗi khi xử lý file Excel: {str(e)}")

    def export_batch_results(self,
                             original_file_path: str,
                             results: List[Dict[str, Any]],
                             output_file_path: str,
                             so_an: int) -> str:
        """
        Xuất kết quả xử lý hàng loạt ra file Excel mới

        Args:
            original_file_path: Đường dẫn file gốc
            results: Danh sách kết quả xử lý
            output_file_path: Đường dẫn file output
            so_an: Số ẩn

        Returns:
            Đường dẫn file đã xuất
        """
        try:
            # Đọc file gốc
            original_df = self.read_excel_data(original_file_path)

            # Tạo DataFrame mới từ kết quả
            output_data = []

            for result in results:
                row_index = result['row_index']
                if row_index >= len(original_df):
                    continue

                # Lấy dòng gốc
                original_row = original_df.iloc[row_index].to_dict()

                # Thêm kết quả xử lý
                output_row = {
                    **original_row,
                    'Keylog_Ma_Hoa': result.get('ket_qua_ma_hoa', ''),
                    'Nghiem_He_Phuong_Trinh': result.get('ket_qua_nghiem', ''),
                    'Ket_Qua_Tong': result.get('ket_qua_tong', ''),
                    'Trang_Thai_Xu_Ly': result.get('trang_thai', 'Thành công'),
                    'Ghi_Chu': result.get('ghi_chu', '')
                }

                output_data.append(output_row)

            # Tạo DataFrame output
            output_df = pd.DataFrame(output_data)

            # Đảm bảo thư mục tồn tại
            os.makedirs(os.path.dirname(output_file_path), exist_ok=True)

            # Xuất file Excel với định dạng
            with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                output_df.to_excel(writer, index=False, sheet_name='Kết Quả Xử Lý Hàng Loạt')

                # Định dạng worksheet
                worksheet = writer.sheets['Kết Quả Xử Lý Hàng Loạt']
                self._format_batch_results_worksheet(worksheet, output_df)

            return output_file_path

        except Exception as e:
            raise Exception(f"Không thể xuất file kết quả: {str(e)}")

    def _format_batch_results_worksheet(self, worksheet, df):
        """Định dạng worksheet cho kết quả hàng loạt"""
        try:
            # Định dạng font cho tiêu đề
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = openpyxl.styles.PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')

            # Định dạng font cho dữ liệu
            data_font = Font(name='Arial', size=10)

            # Định dạng cột kết quả
            result_font = Font(name='Arial', size=10, bold=True, color='2E7D32')

            # Áp dụng định dạng cho tiêu đề
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill

            # Áp dụng định dạng cho dữ liệu
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = data_font

                    # Định dạng đặc biệt cho cột kết quả
                    col_letter = get_column_letter(col)
                    col_name = df.columns[col - 1] if col - 1 < len(df.columns) else ""

                    if any(keyword in col_name for keyword in ['Keylog', 'Nghiem', 'Ket_Qua']):
                        cell.font = result_font

            # Tự động điều chỉnh độ rộng cột
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        except Exception as e:
            print(f"Lỗi định dạng worksheet: {e}")

    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """
        Lấy thông tin về file Excel

        Returns:
            Dictionary chứa thông tin file
        """
        try:
            df = self.read_excel_data(file_path)
            file_name = os.path.basename(file_path)

            return {
                'file_name': file_name,
                'total_rows': len(df),
                'total_columns': len(df.columns),
                'columns': list(df.columns),
                'file_size': os.path.getsize(file_path),
                'first_few_rows': df.head(3).to_dict('records')
            }
        except Exception as e:
            raise Exception(f"Không thể đọc thông tin file: {str(e)}")

    def validate_equation_data_quality(self, df: pd.DataFrame, so_an: int) -> Dict[str, Any]:
        """
        Kiểm tra chất lượng dữ liệu phương trình trong DataFrame

        Returns:
            Dictionary chứa thông tin chất lượng dữ liệu
        """
        key = f"{so_an}_an"
        if key not in self.mapping['equation_mapping_by_phuong_trinh']:
            return {'valid': False, 'error': f"Không hỗ trợ hệ {so_an} ẩn"}

        required_cols = self.mapping['equation_mapping_by_phuong_trinh'][key]['required_columns']

        quality_info = {
            'valid': True,
            'total_rows': len(df),
            'rows_with_data': 0,
            'rows_with_errors': 0,
            'missing_columns': [],
            'data_issues': []
        }

        # Kiểm tra cột thiếu
        for col in required_cols:
            if col not in df.columns:
                quality_info['missing_columns'].append(col)
                quality_info['valid'] = False

        if not quality_info['valid']:
            return quality_info

        # Kiểm tra từng dòng
        for row_index in range(len(df)):
            has_data = False
            has_issue = False
            issues = []

            for col in required_cols:
                value = df.iloc[row_index][col]

                if pd.isna(value) or str(value).strip() == "":
                    issues.append(f"Cột '{col}' trống")
                    has_issue = True
                else:
                    # Kiểm tra định dạng hệ số
                    try:
                        coefficients = str(value).strip().split(',')
                        expected_coeffs = so_an + 1  # Số hệ số mong đợi

                        if len(coefficients) < expected_coeffs:
                            issues.append(f"Cột '{col}' thiếu hệ số (cần {expected_coeffs}, có {len(coefficients)})")
                            has_issue = True
                        elif len(coefficients) > expected_coeffs:
                            issues.append(f"Cột '{col}' thừa hệ số (cần {expected_coeffs}, có {len(coefficients)})")
                            has_issue = True

                        # Kiểm tra hệ số hợp lệ
                        for coeff in coefficients[:expected_coeffs]:
                            coeff = coeff.strip()
                            if coeff and not self._is_valid_coefficient(coeff):
                                issues.append(f"Hệ số không hợp lệ: '{coeff}' trong cột '{col}'")
                                has_issue = True

                    except Exception as e:
                        issues.append(f"Lỗi xử lý dữ liệu cột '{col}': {str(e)}")
                        has_issue = True

                    has_data = True

            if has_data:
                quality_info['rows_with_data'] += 1

            if has_issue:
                quality_info['rows_with_errors'] += 1
                quality_info['data_issues'].append({
                    'row': row_index + 2,  # +2 vì Excel bắt đầu từ 1 và có header
                    'issues': issues
                })

        return quality_info

    def _is_valid_coefficient(self, coeff: str) -> bool:
        """
        Kiểm tra hệ số có hợp lệ không
        Hỗ trợ: số, phân số, biểu thức toán học cơ bản
        """
        coeff = coeff.strip()

        # Cho phép rỗng (sẽ được thay bằng 0)
        if not coeff:
            return True

        # Kiểm tra số
        try:
            # Thử chuyển thành số
            if '/' in coeff:
                # Xử lý phân số
                parts = coeff.split('/')
                if len(parts) == 2:
                    float(parts[0]) / float(parts[1])
                    return True
            else:
                float(coeff)
                return True
        except:
            pass

        # Kiểm tra biểu thức toán học cơ bản
        math_patterns = [
            r'^[0-9+\-*/().\s]+$',  # Biểu thức số học cơ bản
            r'sqrt\([^)]+\)',  # Căn bậc hai
            r'sin\([^)]+\)',  # Sin
            r'cos\([^)]+\)',  # Cos
            r'tan\([^)]+\)',  # Tan
            r'pi',  # Số pi
            r'e',  # Số e
            r'log\([^)]+\)',  # Logarit
            r'ln\([^)]+\)',  # Logarit tự nhiên
        ]

        for pattern in math_patterns:
            if re.search(pattern, coeff, re.IGNORECASE):
                return True

        return False

    def create_equation_template(self, so_an: int, output_path: str) -> str:
        """
        Tạo template Excel cho hệ phương trình

        Args:
            so_an: Số ẩn
            output_path: Đường dẫn xuất template

        Returns:
            Đường dẫn file template đã tạo
        """
        try:
            key = f"{so_an}_an"
            if key not in self.mapping['equation_mapping_by_phuong_trinh']:
                raise Exception(f"Không hỗ trợ template cho hệ {so_an} ẩn")

            required_cols = self.mapping['equation_mapping_by_phuong_trinh'][key]['required_columns']

            # Tạo dữ liệu mẫu
            template_data = {}
            for col in required_cols:
                if so_an == 2:
                    template_data[col] = ['2,1,5', '1,-2,3', '0.5,2,4.5', '3,4,2']
                elif so_an == 3:
                    template_data[col] = ['1,2,3,6', '2,1,-1,8', '4,5,6,15']
                elif so_an == 4:
                    template_data[col] = ['1,0,0,0,1', '2,1,1,0,8', '0,1,0,0,2']

            df = pd.DataFrame(template_data)

            # Thêm dòng hướng dẫn
            description = f"Hệ {so_an} phương trình {so_an} ẩn - Mỗi ô chứa {so_an + 1} hệ số phân cách bằng dấu phẩy"
            instruction_df = pd.DataFrame([{col: description if i == 0 else "" for i, col in enumerate(required_cols)}])
            df = pd.concat([instruction_df, df], ignore_index=True)

            # Xuất template
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name=f'Hệ {so_an} ẩn')

                # Định dạng
                worksheet = writer.sheets[f'Hệ {so_an} ẩn']
                for col in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            return output_path

        except Exception as e:
            raise Exception(f"Không thể tạo template: {str(e)}")

    def export_results(self, original_df: pd.DataFrame, encoded_results: List[str], output_path: str) -> str:
        """Export results với định dạng font cho cột keylog"""
        try:
            result_df = original_df.copy()

            # Tìm cột keylog
            keylog_column = None
            for col in result_df.columns:
                if col.strip().lower() == 'keylog':
                    keylog_column = col
                    break

            # Ghi kết quả
            if keylog_column:
                result_df[keylog_column] = encoded_results
            else:
                result_df['keylog'] = encoded_results

            # Đảm bảo thư mục tồn tại
            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            # Ghi file Excel với định dạng
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False, sheet_name='Results')

                # Định dạng cột keylog
                worksheet = writer.sheets['Results']
                self._format_keylog_column(worksheet, result_df, keylog_column)

            return output_path

        except Exception as e:
            raise Exception(f"Không thể xuất file kết quả: {str(e)}")

    def _format_keylog_column(self, worksheet, df, keylog_column=None):
        """Định dạng font chữ cho cột keylog"""
        try:
            # Xác định tên cột keylog
            keylog_col_name = keylog_column if keylog_column else 'keylog'

            # Tìm vị trí cột
            keylog_col_idx = None
            for idx, col_name in enumerate(df.columns):
                if col_name == keylog_col_name:
                    keylog_col_idx = idx
                    break

            if keylog_col_idx is not None:
                col_letter = get_column_letter(keylog_col_idx + 1)

                # Font cho cột keylog
                keylog_font = Font(
                    name='Arial',
                    size=10,
                    bold=False,
                    color='000000'
                )

                # Áp dụng font cho dữ liệu
                for row in range(2, len(df) + 2):
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.font = keylog_font

                # Điều chỉnh độ rộng cột
                worksheet.column_dimensions[col_letter].width = 50

        except Exception as e:
            print(f"Lỗi định dạng cột keylog: {e}")

    def read_excel_data_chunked(self, file_path: str, chunksize: int = 1000):
        """Đọc file Excel theo từng chunk để xử lý file lớn"""
        try:
            chunk_iterator = pd.read_excel(file_path, chunksize=chunksize)
            return chunk_iterator
        except Exception as e:
            raise Exception(f"Không thể đọc file Excel theo chunk: {str(e)}")

    def get_total_rows(self, file_path: str) -> int:
        """Lấy tổng số dòng để tính tiến độ"""
        try:
            df = pd.read_excel(file_path)
            return len(df)
        except:
            return 0

    def debug_extract_single_row(self, df: pd.DataFrame, so_an: int, row_index: int):
        """
        Hàm debug để kiểm tra trích xuất từng dòng
        """
        key = f"{so_an}_an"
        if key not in self.mapping['equation_mapping_by_phuong_trinh']:
            print(f"Không tìm thấy cấu hình cho hệ {so_an} ẩn")
            return

        columns_config = self.mapping['equation_mapping_by_phuong_trinh'][key]['columns']
        num_coeff_per_eq = so_an + 1

        print(f"=== DEBUG Dòng {row_index}, hệ {so_an} ẩn ===")

        all_coefficients = []

        for col_name, col_config in columns_config.items():
            excel_column = col_config['excel_column']
            value = df.iloc[row_index][excel_column] if excel_column in df.columns else ""
            print(f"  {col_name} ({excel_column}): '{value}'")

            # Xử lý giống hàm chính
            coefficients = []
            if excel_column in df.columns:
                value = df.iloc[row_index][excel_column]
                if pd.isna(value):
                    value = ""
                coefficients = [coeff.strip() for coeff in str(value).strip().split(',') if coeff.strip() != ""]

            # Bổ sung số 0 cho đủ mỗi phương trình
            while len(coefficients) < num_coeff_per_eq:
                coefficients.append("0")
            coefficients = coefficients[:num_coeff_per_eq]

            print(f"    → Hệ số đã xử lý: {coefficients}")
            all_coefficients.extend(coefficients)

        print(f"  Tổng hệ số: {all_coefficients}")
        print(f"  Số hệ số: {len(all_coefficients)}")
        print("=" * 40)